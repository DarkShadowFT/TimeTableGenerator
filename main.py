import time, math, os
import openpyxl
from xlwings import Range, Book
from Period import Period
import pandas as pd
import xlsxwriter, shutil, excel2img


def extractDuration(period):
	return period.duration

def extractDay(period):
	index = weekdays.index(period.day.lower())
	return index

def deleteTxtFiles():
	if os.path.exists('tt.txt'):
		os.remove('tt.txt')
	if os.path.exists('Output.txt'):
		os.remove('Output.txt')
	if os.path.exists('periods.txt'):
		os.remove('periods.txt')

def latestWorkBook():
	newTT = True
	tt_files = []
	for root, dir, files in os.walk("C:\Data\SHIT-NUCES\Current Semester"):
		for file in files:
			if file.lower().find("fast school of computing time table") == 0 and file.lower().find("~$") == -1:
				tt_files.append(os.path.join(root, file))

	with open("tt version.txt", "a") as ttv1:
		pass

	tt_version = ""
	with open("tt version.txt", "r") as ttv2:
		tt_version = ttv2.read()

	if tt_version == "" or (tt_version != '' and tt_version != tt_files[-1]):
		with open("tt version.txt", "w") as ttv3:
			ttv3.write(tt_files[-1])
		deleteTxtFiles()
	else:
		newTT = False
	return newTT

def get_time(start_time, minutes):
	res = ""
	temp = start_time.split(':')
	temp[1] = int(temp[1]) + int(minutes)
	hours = 0
	if temp[1] >= 60:
		hours = temp[1] / 60
		temp[1] %= 60
	elif temp[1] < 0:
		hours = math.floor(temp[1] / 60)
		temp[1] = 60 - (abs(temp[1]) % 60)
	temp[0] = int(temp[0]) + int(hours)
	if temp[0] < 10:
		temp[0] = "0" + str(temp[0])
	if temp[1] < 10:
		temp[1] = "0" + str(temp[1])
	res = str(temp[0]) + ":" + str(temp[1])
	return res

def make_timetable():
	day = ""
	venue = ""
	day_found = False
	none_count = 0
	none_length = 0
	start_time = 0
	start_time_set = False
	p_file = open("periods.txt", "w")

	for i in range(0, len(data)):
		j = 0
		while j < len(data[i]):
			if isinstance(data[i][j], str) and data[i][j].lower().find("period") != -1:
				j += 1
				time_count = 0
				times = []
				for k in range(j, len(data[i])):
					if time_count == 2:
						break
					elif data[i][k] != "None":
						times.insert(k, data[i][k])
						time_count += 1

				none_length = int(times[1]) - int(times[0])

			elif isinstance(data[i][j], str) and data[i][j].lower().find("a.m.") != -1 and not start_time_set:
				start_time_set = True
				start_time = data[i][j].split()[0]

			elif data[i][j] != "None":
				if j == 1 and day_found:
					none_count = 0
					venue = data[i][j]

				elif j > 1 and day_found:
					length = float(none_count) * float(none_length)
					period = data[i][j]
					if period != "NOT AVAILABLE":
						period = period.split('(')
						if len(period) > 1:
							section = period[1].split(')')
							cell_span = 1
							if Range((i + 1, j + 1)).merge_cells:
								cell_span = Range((i + 1, j + 1)).merge_area.count

							period_start_time = get_time(start_time, length)
							period_end_time = get_time(start_time, length + cell_span * none_length)
							p = Period(period_start_time, period_end_time, period[0], section[0], venue, day[0])
							none_count += cell_span
							j += cell_span - 1
							print(p, file = p_file)
				else:
					for k in range(0, len(weekdays)):
						if isinstance(data[i][j], str) and data[i][j].lower().find(weekdays[k]) != -1:
							none_count = 0
							day_found = True
							day = data[i][j].split()
							temp_day = day[0]
							while temp_day.lower() in weekdays:
								j += 1
								temp_day = data[i][j]

							venue = data[i][j]
							break
			else:
				none_count += 1
			j += 1
	
	p_file.close()

if not latestWorkBook():
	print("No new timetable detected")
	exit()


path = ""
with open("tt version.txt", "r") as ttv:
	path = ttv.read()
wb = Book(path)

wb = wb.sheets[0]
wb_obj = openpyxl.load_workbook(path, True)
sheet_obj = wb_obj['FSC TT (Spring 2022)']

print("Workbook loaded")
max_row = sheet_obj.max_row
max_col = sheet_obj.max_column
data = []

# Create file if not created
f1 = open("Output.txt", "a+")
f1.close()

f2 = open("Output.txt", "r")
contents = f2.read()
if len(contents) == 0:
	data = [ ["" for i in range(max_col)] for j in range(max_row) ]

	for i, row in enumerate(sheet_obj.rows):
		for j, cell in enumerate(row):
			data[i][j] = cell.value

	# Removing empty rows
	i = 0
	while i < len(data):
		for j in range(0, len(data[i])):
			if isinstance(data[i][j], str) and data[i][j].lower().find('=count') != -1:
				del data[i]
				break
		i += 1
		
	# Writing the parsed output to "Output.txt"
	with open("Output.txt", "w") as f:
		for i in range(0, len(data)):
			for j in range(0, len(data[i])):
				print(data[i][j], end = '', file = f)
				print('\t', end = '', file = f)
			print(file = f)

with open("Output.txt", "r") as w:
	data = w.read().split('\n')
	for i, words in enumerate(data):
		data[i] = words.split('\t')


weekdays = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]

# Create periods.txt if not created
p1 = open("periods.txt", "a")
p1.close()

periods = []
p2 = open("periods.txt", "r")
p2_contents = p2.read()
if len(p2_contents) == 0:
	make_timetable()
else:
	periods = p2_contents.split('\n')

wb_obj.close()

mon_periods, tue_periods, wed_periods, thu_periods, fri_periods, sat_periods = [], [], [], [], [], []
sections = ["BCS-6A", "BCS-6B", "BCS-6C", "BCS-6D", "BCS-6E", "BCS-6F", "BCS-6G", "BCS-6H", "BCS-6J"]
with open("periods.txt", "r") as w:
	f = w.read().split('\n')
	for period in f:
		if period != '':
			p = period.split('\t\t')
			name = p[0].strip()
			section = p[1]
			startTime = p[3].split('-')[0]
			endTime = p[3].split('-')[1]
			venue = p[2]
			day = p[4]
			period = Period(startTime, endTime, name, section, venue, day)

			if period.day == "Monday" and period.section in sections:
				mon_periods.append(period)
			elif period.day == "Tuesday" and period.section in sections:
				tue_periods.append(period)
			elif period.day == "Wednesday" and period.section in sections:
				wed_periods.append(period)
			elif period.day == "Thursday" and period.section in sections:
				thu_periods.append(period)
			elif period.day == "Friday" and period.section in sections:
				fri_periods.append(period)
			elif period.day == "Saturday" and period.section in sections:
				sat_periods.append(period)

all_day_periods = [mon_periods, tue_periods, wed_periods, thu_periods, fri_periods, sat_periods]

sec_periods = []
for section in sections:
	sec_periods.append([])

with open("tt.txt", "w") as tt:
	for i, section in enumerate(sections):
		temp = f"\n\n\t\t{section}\n\n" 
		tt.write(temp)
		for day in all_day_periods:
			temp = f"\n{day[0].day}\n"
			tt.write(temp)
			for period in day:
				if period.section == section:
					print(period, file = tt)
					sec_periods[sections.index(section)].append(period)


# # Create a Pandas Excel writer using XlsxWriter as the engine.
# writer = pd.ExcelWriter('timetable.xlsx', engine='xlsxwriter')

row_numbers = []

# with open("tt.txt", "w") as tt:
# 	idx = 0
# 	for i, section in enumerate(section_periods):
# 		row_numbers.append(idx)

# 		section_df = pd.DataFrame(["", sections[i], ""]).transpose()
# 		# section_df.to_excel(writer, sheet_name = 'Sheet1', startrow = idx, index = False, header = False)
# 		print(f"\t\t{sections[i]}\n", file = tt)

# 		df = pd.DataFrame([ [s.day, s.name, s.duration, s.venue] for s in section], columns = ["day", "name", "duration", "venue"])

# 		# Convert the dataframe to an XlsxWriter Excel object.
# 		# df.to_excel(writer, sheet_name = 'Sheet1', startrow = idx + 1, index = False, header = False)

# 		idx += len(section) + 2
# 		print(f"{df.to_string(index = False)}\n", file = tt)

# # Close the Pandas Excel writer and output the Excel file.
# writer.save()

workbook = xlsxwriter.Workbook('timetable all sections in Batch.xlsx')
ws1 = workbook.add_worksheet()
ws1.set_column(1, 1, 30)
ws1.set_column(2, 2, 12)
ws1.set_column(3, 3, 11)
merge_format = workbook.add_format({'align': 'center'})
merge_format.set_border()
border_format = workbook.add_format()
border_format.set_border()
with open("tt.txt", "r") as f_tt:
	contents = f_tt.read().split('\n')
	i = 0
	for row in contents:
		if row.strip() == '':
			continue
		period = row.split('\t\t')
		if row.strip() in sections:				
			ws1.write(i, 1, ' ')
			ws1.merge_range(i + 1, 1, i + 1, 3, row.strip(), merge_format)
			i += 1
		elif row.strip().lower() in weekdays:
			ws1.merge_range(i, 1, i, 3, row.strip(), merge_format)
		elif len(period) == 5:
			j = 0
			k = j + 1
			while j < len(period):
				if period[j] not in sections and period[j].lower() not in weekdays:
					ws1.write(i, k, str(period[j]), border_format)
					k += 1
				j += 1
		i += 1
workbook.close()

my_choices = [["Parallel & Dist Computing", "BCS-6C"], ["Compiler Construction", "BCS-6A"], ["Organizational Behaviour", "BCS-6A"], 
				["Artificial Intelligence", "BCS-6B"], ["Software Engineering", "BCS-6F"], ["Artificial Intelligence Lab", "BCS-6B"]]
my_courses = ["Parallel & Dist Computing", "Compiler Construction", "Organizational Behaviour", 
				"Artificial Intelligence", "Software Engineering", "Artificial Intelligence Lab"]
my_courses_short = ["PDC", "CC", "OB", "AI", "SE", "AI Lab"]				  
				
my_periods = []
for section in sec_periods:
	for period in section:
		for my_choice in my_choices:
			if my_choice[0] == period.name and my_choice[1] == period.section:
				my_periods.append(period)

my_period_list = []
for weekday in weekdays:
	my_period_list.append([])

for my_period in my_periods:
	my_period_list[weekdays.index(my_period.day.lower())].append(my_period)

workbook = xlsxwriter.Workbook('timetable.xlsx')
ws1 = workbook.add_worksheet()
merge_format = workbook.add_format({'align': 'center'})
merge_format.set_bg_color('#BDD7EE')

j = 2
k = 0
color_codes = ['#BDD7EE', '#F4B084', '#A9D08E', '#AEAAAA', '#FFD966', '#FF99FF']
ws1.merge_range(j - 1, 1, j - 1, 3, "Our TimeTable", merge_format)
ws1.set_column(1, 3, len(my_courses_short[-1]) + len(sections[0]) + 1)	
ws1.set_zoom(159)
for i, day in enumerate(my_period_list):
	if len(my_period_list[i]) > 0:
		cell_format = workbook.add_format({'align': 'center'})
		cell_format.set_bg_color(color_codes[k])
		cell_format.set_align('center')
		ws1.merge_range(j, 1, j, 3, weekdays[i].upper()[0] + weekdays[i][1:], cell_format)
		j += 1
		# print(f"Day = {weekdays[i]}")
		my_period_list[i] = sorted(my_period_list[i], key=extractDuration)
		day = my_period_list[i]
		for my_period in day:
			cell_format2 = workbook.add_format()
			cell_format2.set_bg_color(color_codes[k])
			# my_period.tt_format()
			ws1.write(j, 1, f"{my_courses_short[my_courses.index(my_period.name)]}({my_period.section})", cell_format2)
			ws1.write(j, 2, str(my_period.venue), cell_format2)
			ws1.write(j, 3, str(my_period.duration), cell_format2)
			j += 1
		k += 1
workbook.close()

excel2img.export_img("timetable.xlsx", "timetable.png", "Sheet1", None)
dest = shutil.copyfile('C:\Data\SHIT-NUCES\Current Semester\TimeTable.png', 'C:\Data\SHIT-NUCES\Current Semester\TimeTable(Copy).png')
dest = shutil.copyfile('timetable.png', 'C:\Data\SHIT-NUCES\Current Semester\TimeTable.png')
dest = shutil.copyfile('C:\Data\SHIT-NUCES\Current Semester\TimeTable.xlsx', 'C:\Data\SHIT-NUCES\Current Semester\TimeTable(Copy).xlsx')
dest = shutil.copyfile('timetable.xlsx', 'C:\Data\SHIT-NUCES\Current Semester\TimeTable.xlsx')
print(f"TimeTable copied at {dest}")

f2.close()